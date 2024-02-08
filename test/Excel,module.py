
listmot = []

import openpyxl
nom_fichier_excel = ("LexiqueNew.xlsx")
nom_feuille_excel = ("fich1")
numero_colonne_excel = (1)
def lire_colonne_excel(nom_fichier, nom_feuille, numero_colonne):
    """
    Lire les valeurs d'une colonne spécifique dans un fichier Excel.

    Args:
    - nom_fichier (str): Le nom du fichier Excel.
    - nom_feuille (str): Le nom de la feuille Excel.
    - numero_colonne (int): Le numéro de la colonne à lire (index de 1).

    Returns:
    - list: Une liste des valeurs de la colonne spécifiée.
    """
    valeurs_colonne = []

    # Charger le classeur Excel
    workbook = openpyxl.load_workbook(nom_fichier)

    try:
        # Accéder à la feuille
        sheet = workbook[nom_feuille]

        # Lire les valeurs de la colonne spécifiée
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=numero_colonne, max_col=numero_colonne):
            for cell in row:
                valeurs_colonne.append(cell.value)
        print(valeurs_colonne[1])
    finally:
        # Fermer le classeur Excel
        workbook.close()

    return valeurs_colonne
lire_colonne_excel("LexiqueNew.xlsx", "fich1", 1)